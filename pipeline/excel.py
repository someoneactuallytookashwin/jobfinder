"""Excel read/write for results.xlsx.

Creates the workbook with headers on first run, appends rows on later runs,
highlights the top 10 ranked rows green and any "tailored" rows blue, and
exposes a helper to update a job's status/resume-file after tailoring.

YCombinator jobs are written to their own "YCombinator" worksheet so they can
be reviewed separately; every other source goes to the main "Jobs" sheet.
"""

from __future__ import annotations

import os

import config
from pipeline.utils import get_logger, now_iso

logger = get_logger()

MAIN_SHEET = "Jobs"
YC_SHEET = "YCombinator"
YC_SOURCE = "ycombinator"

COLUMNS = [
    "Run Date",
    "Rank",
    "Score",
    "Title",
    "Company",
    "Location",
    "Source",
    "Posted",
    "URL",
    "Score Reasons",
    "Missing Skills",
    "ATS Keywords",
    "Status",
    "Resume File",
    "Job ID",  # last column — used to locate rows for status updates
]

_GREEN = "C6EFCE"   # top-10 highlight
_BLUE = "BDD7EE"    # tailored highlight
_HEADER = "4472C4"  # header fill


def _styles():
    from openpyxl.styles import Font, PatternFill

    return {
        "green": PatternFill(start_color=_GREEN, end_color=_GREEN, fill_type="solid"),
        "blue": PatternFill(start_color=_BLUE, end_color=_BLUE, fill_type="solid"),
        "header_fill": PatternFill(start_color=_HEADER, end_color=_HEADER, fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF"),
    }


def _job_to_row(job: dict, run_date: str) -> list:
    return [
        run_date,
        job.get("rank", ""),
        job.get("score", ""),
        job.get("title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("source", ""),
        job.get("posted_at", ""),
        job.get("url", ""),
        job.get("score_reason", ""),
        job.get("missing", ""),
        job.get("ats_keywords", ""),
        job.get("status", ""),
        job.get("resume_file", "") or "",
        job.get("id", ""),
    ]


def _load_workbook():
    """Load the existing workbook, or create a fresh one (no default sheet)."""
    from openpyxl import Workbook, load_workbook

    if os.path.exists(config.EXCEL_PATH):
        return load_workbook(config.EXCEL_PATH)

    wb = Workbook()
    # Drop the auto-created default sheet; we add named sheets on demand.
    wb.remove(wb.active)
    return wb


def _get_sheet(wb, name: str):
    """Return the named worksheet, creating it (with styled headers) if absent."""
    if name in wb.sheetnames:
        return wb[name]

    ws = wb.create_sheet(title=name)
    styles = _styles()
    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
    ws.freeze_panes = "A2"
    return ws


def write_results(jobs: list[dict]) -> None:
    """Append scored jobs as new rows (does not overwrite prior runs).

    Routes YCombinator jobs to the YC_SHEET and everything else to MAIN_SHEET.
    """
    os.makedirs(config.DATA_DIR, exist_ok=True)
    wb = _load_workbook()
    styles = _styles()
    run_date = now_iso()

    counts = {MAIN_SHEET: 0, YC_SHEET: 0}
    for job in jobs:
        sheet_name = YC_SHEET if job.get("source") == YC_SOURCE else MAIN_SHEET
        ws = _get_sheet(wb, sheet_name)
        ws.append(_job_to_row(job, run_date))
        counts[sheet_name] += 1

        row_idx = ws.max_row
        rank = job.get("rank") or 0
        status = (job.get("status") or "").lower()

        fill = None
        if status == "tailored":
            fill = styles["blue"]
        elif isinstance(rank, int) and 1 <= rank <= 10:
            fill = styles["green"]

        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Keep the main sheet first in tab order for convenience.
    _order_sheets(wb)
    for name in (MAIN_SHEET, YC_SHEET):
        if name in wb.sheetnames:
            _autosize(wb[name])
    wb.save(config.EXCEL_PATH)
    logger.info(
        "Wrote %d rows to %s (Jobs: %d, YCombinator: %d).",
        len(jobs), config.EXCEL_PATH, counts[MAIN_SHEET], counts[YC_SHEET],
    )


def _order_sheets(wb) -> None:
    """Ensure MAIN_SHEET appears before YC_SHEET in the tab bar."""
    desired = [n for n in (MAIN_SHEET, YC_SHEET) if n in wb.sheetnames]
    desired += [n for n in wb.sheetnames if n not in desired]
    wb._sheets.sort(key=lambda s: desired.index(s.title))


def update_job_status(job_id: str, *, status: str, resume_file: str | None = None) -> bool:
    """Update the most recent row matching job_id across BOTH sheets (Status /
    Resume File columns) and re-apply the tailored (blue) highlight. Returns True
    if a row updated.
    """
    if not os.path.exists(config.EXCEL_PATH):
        logger.warning("No results.xlsx to update for job %s.", job_id)
        return False

    from openpyxl import load_workbook

    wb = load_workbook(config.EXCEL_PATH)
    styles = _styles()

    status_col = COLUMNS.index("Status") + 1
    resume_col = COLUMNS.index("Resume File") + 1
    id_col = COLUMNS.index("Job ID") + 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(ws.max_row, 1, -1):  # search from the bottom (latest run)
            if str(ws.cell(row=row, column=id_col).value) != str(job_id):
                continue
            ws.cell(row=row, column=status_col).value = status
            if resume_file is not None:
                ws.cell(row=row, column=resume_col).value = resume_file
            if status.lower() == "tailored":
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = styles["blue"]
            wb.save(config.EXCEL_PATH)
            logger.info("Updated job %s → status=%s (%s sheet).", job_id, status, sheet_name)
            return True

    logger.warning("Job id %s not found in results.xlsx.", job_id)
    return False


def _autosize(ws) -> None:
    """Roughly size columns to their content (capped) for readability."""
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, len(COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(COLUMNS[col_idx - 1]))
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
